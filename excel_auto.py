import time
import threading
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

flag = True


# function to count time how long it takes to process the file
def waiting():
    global flag
    count = 0
    while flag:
        print('Working.')
        print(f'Time: {count} s')
        count += 1
        time.sleep(1)
        print('Working..')
        print(f'Time: {count} s')
        count += 1
        time.sleep(1)
        print('Working...')
        print(f'Time: {count} s')
        count += 1
        time.sleep(1)

    print(f'Ended within {count} s')


# function that process the file
def sheet():
    # global flag, when loops ends it's set to False and end waiting func
    global flag
    # load workbook // source: https://catalog.data.gov/dataset/electric-vehicle-population-data
    wb = load_workbook('Electric_Vehicle_Population_Data.xlsx')
    # set acctual sheet to 'input_data'
    ws = wb['input_data']
    # check if there is sheet names 'output_data' if not, make it, if yes pass to avoid make new every program run
    if 'output_data' in wb.sheetnames:
        pass
    else:
        wb.create_sheet('output_data')

    # loop that check all used rows in sheet, take values and split it into list
    for i in range(1, ws.max_row + 1):
        # make sure we work in input_data sheet
        ws = wb['input_data']
        raw_data = ws[f'A{i}'].value
        list_raw_data = raw_data.split(',')
        # set output_data as working sheet
        ws = wb['output_data']
        # loop that now takes all elements from list and put in consecutive cells
        for j in range(len(list_raw_data)):
            char = get_column_letter(j + 1)
            ws[f'{char}{i}'].value = list_raw_data[j]

    # make sure we work in proper sheet
    ws = wb['output_data']
    # now we loop through every value in column I to check if vehicle is BEV or PHEV and colour cells
    for i in ws['I']:
        if i.value == 'Battery Electric Vehicle (BEV)':
            i.fill = PatternFill(start_color='38761d', end_color='38761d', fill_type='solid')
        elif i.value == 'Plug-in Hybrid Electric Vehicle (PHEV)':
            i.fill = PatternFill(start_color='8fce00', end_color='8fce00', fill_type='solid')
        else:
            pass
    # now we loop through every value in column L and colour cells red if value is 0, and green if it's above 0
    for i in ws['L']:
        if i.value != '0':
            i.fill = PatternFill(start_color='38761d', end_color='38761d', fill_type='solid')
        else:
            i.fill = PatternFill(start_color='bf1e1a', end_color='bf1e1a', fill_type='solid')

    # we save file
    wb.save('Electric_Vehicle_Population_Data.xlsx')
    # now set flag to False to end waiting funct
    flag = False


if __name__ == '__main__':
    t1 = threading.Thread(target=waiting)
    t2 = threading.Thread(target=sheet)

    t1.start()
    t2.start()
